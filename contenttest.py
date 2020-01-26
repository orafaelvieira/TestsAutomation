from selenium import webdriver
from selenium.common.exceptions import ElementNotVisibleException, TimeoutException, NoSuchElementException
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

import time
import csv

driver_firefox = webdriver.Firefox(executable_path=r"/Users/rafael/PycharmProjects/TestsAutomation/drivers/geckodriver")

driver = [driver_firefox]
# site = ["https://tuilte-ibe-funnel.int.silversurfer7.com/details/giata:20520?from=20191224&duration=15&adults=2"]
numbertest = 0
giata = []

with open("/Users/rafael/PycharmProjects/TestsAutomation/mvp1test/goalid_brazil.csv", "r") as giatasfile:
    reader = csv.reader(giatasfile, delimiter=",")
    for row in reader:
        giata.append(row[0])
    # print(giata)

components = [".lodging-name", ]

wb_name = "/Users/rafael/PycharmProjects/TestsAutomation/mvp1test/NewTest.xlsx"
wb = load_workbook(wb_name)
results_sheet = wb.create_sheet("Test Results", 0)

results_sheet.title = "Test Results"
results_sheet.sheet_properties.tabColor = "1072BA"


def save_test(giata, workbook_name, sheet_name, numbertest, hotel, test_result, description_result, amenities_result):
    # spreadsheet titles
    sheet_name.cell(row=1, column=1, value="Test")
    sheet_name.cell(row=1, column=2, value="GIATA ID")
    sheet_name.cell(row=1, column=3, value="Property Name")
    sheet_name.cell(row=1, column=4, value="Name Result")
    sheet_name.cell(row=1, column=5, value="Description Result")
    sheet_name.cell(row=1, column=6, value="Amenities Result")

    # spreadsheet lines
    sheet_name.cell(row=numbertest + 1, column=1, value=numbertest)
    sheet_name.cell(row=numbertest + 1, column=2, value=giata)
    sheet_name.cell(row=numbertest + 1, column=3, value=hotel)
    sheet_name.cell(row=numbertest + 1, column=4, value=test_result)
    sheet_name.cell(row=numbertest + 1, column=5, value=description_result)
    sheet_name.cell(row=numbertest + 1, column=6, value=amenities_result)
    wb.save(workbook_name)

def check_amenities(descriptionresult, numbertest):
    test = "True"
    while test == "True":

        try:
            driver_firefox.find_element_by_link_text("› Show more").text == "› Show more"
            amenities_result = "Found"
            print(
                f"Test: {numbertest} - The {hotelname} on hotel {id}. Description: {descriptionresult}. Amenities: {amenities_result}")
            save_test(id, wb_name, results_sheet, numbertest, hotelname, test_result, description_result,
                      amenities_result)
            break

        except ElementNotVisibleException:
            amenities_result = "Not Found - Not Visible"
            print(
                f"Test: {numbertest} - The {hotelname} on hotel {id}. Description: {descriptionresult}. Amenities: {amenities_result}")
            save_test(id, wb_name, results_sheet, numbertest, hotelname, test_result, description_result,
                      amenities_result)
            break

        except TimeoutException:
            amenities_result = "Not Found - Timeout"
            print(
                f"Test: {numbertest} - The {hotelname} on hotel {id}. Description: {descriptionresult}. Amenities: {amenities_result}")
            save_test(id, wb_name, results_sheet, numbertest, hotelname, test_result, description_result,
                      amenities_result)
            break

        except NoSuchElementException:
            amenities_result = "Not Found - No such element"
            print(
                f"Test: {numbertest} - The {hotelname} on hotel {id}. Description: {descriptionresult}. Amenities: {amenities_result}")
            save_test(id, wb_name, results_sheet, numbertest, hotelname, test_result, description_result,
                      amenities_result)
            break

# TRANSFORMAR O ABAIXO EM UMA FUNCAO, CADA COLUNA E UM TESTE?
# Test cases

for id in giata:
    for component in components:
        # Open page, input a basic query
        try:
            driver_firefox.set_page_load_timeout(10)
            driver_firefox.get(
                f"https://tui-lte:tuiLTE!@tuilte-ibe-funnel.stg.silversurfer7.com/detail/gi:{id}?adults=2&children=&from=20200211&to=20200212")
            time.sleep(2)
            driver_firefox.find_element_by_css_selector(component).click()
            hotelname = driver_firefox.find_element_by_css_selector(component).text
            test_result = "Passed"

            # while test_result == "Passed":
            try:
                driver_firefox.find_element_by_css_selector(".snippet h2").text == "What you can expect"
                description_result = "Description Found"

                while description_result == "Description Found":
                    numbertest += 1
                    check_amenities(description_result, numbertest)
                    break

            except ElementNotVisibleException:
                description_result = "Description Not Found"
                numbertest += 1
                check_amenities(description_result, numbertest)
                break

            except TimeoutException:
                description_result = "Description Not Found"
                numbertest += 1
                check_amenities(description_result, numbertest)
                break

            except NoSuchElementException:
                description_result = "Description Not Found"
                numbertest += 1
                check_amenities(description_result, numbertest)
                break

        except ElementNotVisibleException:
            test_result = "Not Visible"
            hotelname = "Not Found"
            description_result = "N/A"
            numbertest += 1
            check_amenities(description_result, numbertest)

        except TimeoutException:
            test_result = "Timeout"
            description_result = "Description Not Found"
            hotelname = "Not Found"
            numbertest += 1
            check_amenities(description_result, numbertest)

        except NoSuchElementException:
            test_result = "No such element"
            description_result = "Description Not Found"
            hotelname = "Not Found"
            numbertest += 1
            check_amenities(description_result, numbertest)