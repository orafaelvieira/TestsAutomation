from selenium import webdriver
from selenium.common.exceptions import ElementNotVisibleException, TimeoutException, NoSuchElementException
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

import time
import csv

# lists and dictionaries

driver = webdriver.Firefox(executable_path=r"/Users/rafael/PycharmProjects/TestsAutomation/drivers/geckodriver")
numbertest = 0
giata = []
# click_components = [".header-logo > .brand-logo",".lodging-name", ".detail-lodging-address", ".markdown-wrapper > p",".room-list-header > .page-heading"]
click_components = {
    "Logo": ".header-logo > .brand-logo",
    "Hotel Name": ".lodging-name",
    "Address": ".detail-lodging-address",
    "Description": ".markdown-wrapper > p",
    "Room list": ".room-list-header > .page-heading",
    "Localization": ".detail-location-heading",
    "Availability": ".loading-button__button-value",
}

with open("/Users/rafael/PycharmProjects/TestsAutomation/mvp1test/goalid_brazil.csv", "r") as giatas_file:
    reader = csv.reader(giatas_file, delimiter=",")
    for row in reader:
        giata.append(row[0])

# Spreadsheet setup

wb_name = "/Users/rafael/PycharmProjects/TestsAutomation/mvp1test/NewTest.xlsx"
wb = load_workbook(wb_name)
results_sheet = wb.create_sheet("Test Results", 0)
results_sheet.title = "Test Results"
results_sheet.sheet_properties.tabColor = "1072BA"

def click_test(css_element,hoteltrial,trial,log_file):
    try:
        driver.find_element_by_css_selector(css_element).is_displayed()
        # text = driver.find_element_by_css_selector(css_element).text
        test_result = "Huge Success"
        print(f"{hoteltrial}.{trial} - Testing now id: {id} - component: {component}, RESULT: {test_result}")
        results_sheet.cell(row=hoteltrial, column=1, value=id)
        results_sheet.cell(row=hoteltrial, column=trial + 1, value=test_result)
        wb.save(log_file)

    except ElementNotVisibleException:
        test_result = "Not Visible"
        print(f"{hoteltrial}.{trial} - Testing now id: {id} - component: {component}, RESULT: {test_result}")
        results_sheet.cell(row=hoteltrial, column=1, value=id)
        results_sheet.cell(row=hoteltrial, column=trial + 1, value=test_result)
        wb.save(log_file)

    except TimeoutException:
        test_result = "Timeout"
        print(f"{hoteltrial}.{trial} - Testing now id: {id} - component: {component}, RESULT: {test_result}")
        results_sheet.cell(row=hoteltrial, column=1, value=id)
        results_sheet.cell(row=hoteltrial, column=trial + 1, value=test_result)
        wb.save(log_file)

    except NoSuchElementException:
        test_result = "No such element"
        print(f"{hoteltrial}.{trial} - Testing now id: {id} - component: {component}, RESULT: {test_result}")
        results_sheet.cell(row=hoteltrial, column=1, value=id)
        results_sheet.cell(row=hoteltrial, column=trial + 1, value=test_result)
        wb.save(log_file)

# id = input("What Hotel would you like to test? Please input the GIATA ID:")

line = 1
current_test = 0
column = 1
number_of_tests = 100000

for component in click_components:
    results_sheet.cell(row=1, column=1, value="GIATA ID")
    results_sheet.cell(row=1, column=column + 1, value=component)
    wb.save(wb_name)
    column += 1

print("Aperture Science")
# while current_test != number_of_tests:

for id in giata:
    if current_test < number_of_tests:
        line += 1
        current_test += 1
        hotel_url = f"https://tui-lte:tuiLTE!@tuilte-guidence-br.stg.silversurfer7.com/detail/goal:{id}?from=20200312&to=20200317&adults=2&utm_source=tripadvisor"
        print(f"New test! Hotel {id}, test number = {current_test} - {hotel_url} ")
        driver.set_page_load_timeout(10)
        driver.get(f"https://tui-lte:tuiLTE!@tuilte-guidence-br.stg.silversurfer7.com/detail/goal:{id}?from=20200312&to=20200317&adults=2&utm_source=tripadvisor")

        time.sleep(5)
        column = 0
        for component in click_components:
            column += 1
            click_test(click_components[component],line,column,wb_name)
    else:
        break

