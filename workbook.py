from openpyxl import load_workbook

trials = [2, 3, 4, 5, 6]
colors = ["blue", "red", "yellow", "vsf","cu"]

wb = load_workbook("TestResults.xlsx")
results_sheet = wb.create_sheet("Results", 0)

results_sheet.title = "Test Results"
results_sheet.sheet_properties.tabColor = "1072BA"

results_sheet.cell(row=1,column=1,value="ID")
results_sheet.cell(row=1,column=2,value="trial")
results_sheet.cell(row=1,column=3,value="color")
current_row = 2
for trial in trials:
    for color in colors:
        results_sheet.cell(row=current_row, column=1, value=current_row-1)
        results_sheet.cell(row=current_row, column=2, value=trial)
        results_sheet.cell(row=current_row, column=3, value=color)
        current_row += 1

wb.save("TestResults.xlsx")