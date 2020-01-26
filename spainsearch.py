from selenium import webdriver
from selenium.common.exceptions import ElementNotVisibleException, TimeoutException, NoSuchElementException
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

import time

driver_chrome = webdriver.Chrome("/Users/rafael/PycharmProjects/TestsAutomation/drivers/chromedriver")
driver_firefox = webdriver.Firefox(executable_path=r"/Users/rafael/PycharmProjects/TestsAutomation/drivers/geckodriver")

drivers = [driver_chrome,driver_firefox]
websites = ["https://esp.tui.com/"]
destinations = ["Madrid","Barcelona"]
adult_passengers = ["//option[. = '1 adulto']","//option[. = '2 adultos']","//option[. = '3 adultos']","//option[. = '4 adultos']","//option[. = '5 adultos']"]
child_passengers = ["//option[. = 'Sin ninños']","//option[. = '1 ninño']","//option[. = '2 ninños']","//option[. = '3 ninños']"]
type_of_travel = ["Package","Hotel"]
test_number = 0

wb_name = "NewTest.xlsx"
wb = load_workbook(wb_name)
results_sheet = wb.create_sheet("Test Results", 0)


results_sheet.title = "Test Results"
results_sheet.sheet_properties.tabColor = "1072BA"


def save_test(site, workbook_name, sheet_name, test_number, driver, city, adult, child, test_result):

    # spreadsheet titles
    sheet_name.cell(row=1, column=1, value="ID")
    sheet_name.cell(row=1, column=2, value="Site")
    sheet_name.cell(row=1, column=3, value="Driver")
    sheet_name.cell(row=1, column=4, value="City")
    sheet_name.cell(row=1, column=5, value="Adults")
    sheet_name.cell(row=1, column=6, value="Children")
    sheet_name.cell(row=1, column=7, value="Test Result")

    # spreadsheet lines
    sheet_name.cell(row=test_number + 1, column=1, value=test_number)
    sheet_name.cell(row=test_number + 1, column=2, value=site)
    sheet_name.cell(row=test_number + 1, column=3, value=driver)
    sheet_name.cell(row=test_number + 1, column=4, value=city)
    sheet_name.cell(row=test_number + 1, column=5, value=adult)
    sheet_name.cell(row=test_number + 1, column=6, value=child)
    if test_result == "Passed":
        sheet_name.cell(row=test_number+1, column=7, value=test_result).fill = PatternFill("solid", fgColor="aad700")
        wb.save(workbook_name)
    else:
        sheet_name.cell(row=test_number+1, column=7, value=test_result).fill = PatternFill("solid", fgColor="ef0000")
        wb.save(workbook_name)

## Test cases
for site in websites:
    for driver in drivers:
        for city in destinations:
            for adult in adult_passengers:
                for child in child_passengers:

                    ## Open page, input a basic query
                    try:
                        driver.set_page_load_timeout(10)
                        driver.get(site)
                        ##driver.find_element_by_css_selector(".tabs > a:nth-child(2)").click()
                        ##driver.find_element_by_css_selector(".property-cell-lg:nth-child(3) > .cell-title strong").click()
                        driver.find_element_by_css_selector(".update-placeholder").click()
                        driver.find_element_by_css_selector(".update-placeholder").send_keys(city)
                        time.sleep(1)
                        driver.find_element_by_css_selector("ul:nth-child(4) em").click()
                        driver.find_element_by_css_selector(".property-cell:nth-child(1) small").click()
                        driver.find_element_by_css_selector(
                            "table:nth-child(3) .row:nth-child(4) > td:nth-child(2) > a").click()
                        time.sleep(1)
                        driver.find_element_by_css_selector(
                            "table:nth-child(3) .row:nth-child(4) > td:nth-child(3) > a").click()
                        time.sleep(1)

                        ## defining number of PAX

                        driver.find_element_by_css_selector(".property-cell-truncate > .cell-title strong").click()
                        time.sleep(2)
                        driver.find_element_by_name("Adults").click()
                        time.sleep(2)

                        dropdown = driver.find_element_by_name("Adults")
                        time.sleep(2)
                        dropdown.find_element_by_xpath(adult).click()
                        time.sleep(2)

                        dropdown = driver.find_element_by_name("Children")
                        time.sleep(2)
                        dropdown.find_element_by_xpath(child).click()
                        time.sleep(2)

                        driver.find_element_by_css_selector(".popout-max-width .btn").click()
                        time.sleep(2)

                        driver.find_element_by_css_selector(".btn-cta").click()

                        test_number += 1
                        print(f"Test on {site}, number {test_number}: Using {driver}, {city},{adult} and {child} Passed")
                        test_result = "Passed"
                        save_test(site, wb_name, results_sheet, test_number, str(driver), city, adult, child, test_result)

                    except ElementNotVisibleException:
                        test_number += 1
                        print(f"Test on {site}, number {test_number}: Using {driver}, {city},{adult} and {child} Failed: Not Visible")
                        test_result = "Failed: Not Visible"
                        save_test(site, wb_name, results_sheet, test_number, str(driver), city, adult, child, test_result)
                    except TimeoutException:
                        test_number += 1
                        print(f"Test on {site}, number {test_number}: Using {driver}, {city},{adult} and {child} Failed: Timeout")
                        test_result = "Failed: Timeout"
                        save_test(site, wb_name, results_sheet, test_number, str(driver), city, adult, child, test_result)
                    except NoSuchElementException:
                        test_number += 1
                        print(f"Test on {site}, number {test_number}: Using {driver}, {city},{adult} and {child} Failed: Timeout")
                        test_result = "Failed: No such element"
                        save_test(site, wb_name, results_sheet, test_number, str(driver), city, adult, child, test_result)